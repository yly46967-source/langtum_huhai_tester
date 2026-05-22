"""Parse uploaded txt files and match with Excel equipment metadata."""
import re
import json
from openpyxl import load_workbook
from config import EQUIPMENT_EXCEL


def parse_txt_content(content: str) -> list[dict]:
    """Parse non-standard JSON (unquoted keys) from txt file content."""
    fixed = re.sub(r'^(\s*)(\w+)(:)', r'\1"\2":', content, flags=re.MULTILINE)
    return json.loads(fixed)


def ship_key_from_filename(filename: str) -> str:
    """Extract ship key from filename, stripping any folder path.

    '测试1-5/ARISTA-散货船-40008_设备信息数组.txt' -> 'ARISTA-散货船-40008'
    'ARISTA-散货船-40008_设备信息数组.txt' -> 'ARISTA-散货船-40008'
    """
    base = filename.replace("\\", "/").rsplit("/", 1)[-1]
    return base.replace("_设备信息数组.txt", "").replace(".txt", "")


def _normalize_key(key: str) -> str:
    return key.replace(" ", "_").lower()


def load_equipment_excel() -> dict:
    """Load 20艘船舶设备列表.xlsx -> {normalized_sheet_name: {ship_info, rows}}."""
    wb = load_workbook(EQUIPMENT_EXCEL, read_only=True, data_only=True)
    result = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        ship_info = {}
        if rows:
            ship_info["company"] = rows[0][0] or ""
            ship_info["ship_name"] = rows[0][1] or ""
            ship_info["ship_type"] = rows[0][2] or ""
        equip_rows = []
        for r in rows:
            equip_rows.append({
                "equip_code": r[7] or "",
                "equip_group": r[8] or "",
            })
        result[_normalize_key(ws.title)] = {"ship_info": ship_info, "rows": equip_rows}
    wb.close()
    return result


def parse_uploaded_file(content: str, filename: str, excel_data: dict) -> dict:
    """Parse an uploaded txt file and enrich with Excel metadata.

    Returns: {ship_key, ship_name, ship_type, company, items: [{manufacturer, model, source_name, equip_code, equip_group}]}
    """
    raw_items = parse_txt_content(content)
    ship_key = ship_key_from_filename(filename)

    norm_key = _normalize_key(ship_key)
    sheet_data = excel_data.get(norm_key, {"ship_info": {}, "rows": []})
    ship_info = sheet_data["ship_info"]
    excel_rows = sheet_data["rows"]

    items = []
    for i, item in enumerate(raw_items):
        equip_code = ""
        equip_group = ""
        if i < len(excel_rows):
            equip_code = excel_rows[i]["equip_code"]
            equip_group = excel_rows[i]["equip_group"]
        items.append({
            "manufacturer": item.get("manufacturer", ""),
            "model": item.get("model", ""),
            "source_name": item.get("source_name", ""),
            "equip_code": equip_code,
            "equip_group": equip_group,
        })

    return {
        "ship_key": ship_key,
        "ship_name": ship_info.get("ship_name", ship_key.split("-")[0]),
        "ship_type": ship_info.get("ship_type", ""),
        "company": ship_info.get("company", ""),
        "items": items,
    }
