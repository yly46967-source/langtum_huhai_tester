"""Generate output Excel from workflow results, matching the customer template format."""
import json
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from db import conn_ctx, get_file, get_success_tasks
from config import RESULT_DIR

logger = logging.getLogger(__name__)

OUTPUT_HEADERS = [
    "船名", "船型", "船籍国", "船级社", "IMO编号", "所属公司",
    "设备编码", "设备名称", "设备型号", "制造商",
    "保养类型",
    "保养项目", "保养步骤", "固定周期数值", "固定周期单位",
    "运行时长数值", "运行时长单位", "是否不定期", "是否必须",
    "安全注意事项", "预计工时", "所需备件", "参考依据",
    "参考设备说明书", "参考政策文件",
]

COLUMN_WIDTHS = {
    1: 20, 2: 8, 3: 8, 4: 8, 5: 14, 6: 20,
    7: 14, 8: 20, 9: 16, 10: 16,
    11: 12, 12: 16, 13: 60, 14: 14, 15: 12,
    16: 14, 17: 12, 18: 12, 19: 10,
    20: 40, 21: 10, 22: 16, 23: 60,
    24: 24, 25: 24,
}

# JSON item field -> Excel column index (1-based, columns 11-25)
ITEM_FIELD_MAP = {
    "maintenance_type": 11,       # 保养类型
    "maintenance_item": 12,       # 保养项目
    "recommended_procedure": 13,  # 保养步骤
    "fixed_interval_value": 14,   # 固定周期数值
    "fixed_interval_unit": 15,    # 固定周期单位
    "runtime_threshold_value": 16, # 运行时长数值
    "runtime_threshold_unit": 17,  # 运行时长单位
    "has_non_periodic_requirement": 18,  # 是否不定期
    "is_mandatory": 19,           # 是否必须
    "safety_notes": 20,           # 安全注意事项
    "estimated_work_hours": 21,   # 预计工时
    "recommended_spares": 22,     # 所需备件
    "source_information": 23,     # 参考依据
}

_HEADER_FONT = Font(name="Carlito", size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(patternType="solid", fgColor="4472C4")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_BORDER = Border(
    left=Side("thin", color="CBD5E1"),
    right=Side("thin", color="CBD5E1"),
    top=Side("thin", color="CBD5E1"),
    bottom=Side("thin", color="CBD5E1"),
)

_DATA_FONT = Font(name="Carlito", size=11)
_DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
_DATA_BORDER = Border(
    left=Side("thin", color="E2E8F0"),
    right=Side("thin", color="E2E8F0"),
    top=Side("thin", color="E2E8F0"),
    bottom=Side("thin", color="E2E8F0"),
)

_ROW_HEIGHT = 15


def extract_end_items(raw_output: str) -> list[dict]:
    """Extract end_item list from raw_output JSON stored in DB.

    raw_output = json.dumps(result.get("detail", {}))
    The detail dict has structure: {"end_item": [...]} or may have nested output.
    """
    try:
        data = json.loads(raw_output)
        # detail may be wrapped: {"output": {"end_item": [...]}} or {"end_item": [...]}
        end_items = data.get("end_item")
        if end_items is None and "output" in data:
            end_items = data["output"].get("end_item", [])
        return end_items or []
    except (json.JSONDecodeError, AttributeError):
        return []


def _format_value(val) -> str:
    """Format a JSON value for Excel cell."""
    if val is None:
        return ""
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False)
    return str(val)


def generate_file_excel(file_id: int) -> str:
    """Generate Excel for a file from all successful tasks."""
    os.makedirs(RESULT_DIR, exist_ok=True)

    with conn_ctx() as conn:
        file_rec = get_file(conn, file_id)
        tasks = get_success_tasks(conn, file_id)

    wb = Workbook()
    ws = wb.active
    ws.title = (file_rec.get("ship_name") or file_rec["ship_key"])[:31]

    for col_idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header row
    ws.row_dimensions[1].height = _ROW_HEIGHT
    for col, header in enumerate(OUTPUT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _HEADER_BORDER

    # Data rows
    row_num = 2
    for task in tasks:
        raw = task.get("raw_output") or ""
        end_items = extract_end_items(raw)

        # Ship info columns (1-6): from file record
        ship_cols = [
            file_rec.get("ship_name", ""),
            file_rec.get("ship_type", ""),
            "",  # 船籍国
            "",  # 船级社
            "",  # IMO编号
            file_rec.get("company", ""),
        ]

        if not end_items:
            # No results — write one row with device info from task
            row_vals = ship_cols + [
                task.get("equip_code", ""),   # 设备编码
                task["source_name"],           # 设备名称
                task["model"],                 # 设备型号
                task["manufacturer"],          # 制造商
            ] + [""] * 14  # empty maintenance columns
            _write_data_row(ws, row_num, row_vals)
            row_num += 1
            continue

        # Each end_item corresponds to one equipment from the workflow output
        for ei in end_items:
            # Device info from end_item (workflow output takes precedence)
            ei_source = ei.get("source_name", task["source_name"])
            ei_manufacturer = ei.get("manufacturer", task["manufacturer"])
            ei_model = ei.get("model", task["model"])

            device_cols = [
                task.get("equip_code", ""),  # 设备编码
                ei_source,                    # 设备名称
                ei_model,                     # 设备型号
                ei_manufacturer,              # 制造商
            ]

            items = ei.get("items", [])
            if not items:
                # end_item with no maintenance items — write one row
                row_vals = ship_cols + device_cols + [""] * 14
                _write_data_row(ws, row_num, row_vals)
                row_num += 1
                continue

            # Each item is one maintenance row
            for m_item in items:
                # Build maintenance columns (11-25)
                maint_cols = []
                for _ in range(14):
                    maint_cols.append("")

                for field, col_idx in ITEM_FIELD_MAP.items():
                    val = m_item.get(field, "")
                    excel_col = col_idx - 11  # index into maint_cols
                    maint_cols[excel_col] = _format_value(val)

                # Column 24 (index 13): 参考设备说明书 — from source_information[].file_name
                si_list = m_item.get("source_information", [])
                file_names = list(dict.fromkeys(
                    s.get("file_name") for s in si_list if s.get("file_name")
                ))
                if file_names:
                    maint_cols[13] = ", ".join(file_names)

                row_vals = ship_cols + device_cols + maint_cols
                _write_data_row(ws, row_num, row_vals)
                row_num += 1

    safe_key = file_rec['ship_key'].replace('/', '_').replace('\\', '_')
    output_path = os.path.join(RESULT_DIR, f"{safe_key}_结果.xlsx")
    wb.save(output_path)
    logger.info(f"Generated Excel: {output_path} ({row_num - 1} data rows)")
    return output_path


def _write_data_row(ws, row_num: int, values: list):
    ws.row_dimensions[row_num].height = _ROW_HEIGHT
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = _DATA_FONT
        cell.alignment = _DATA_ALIGN
        cell.border = _DATA_BORDER
